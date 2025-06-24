from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages

from .models import Estudiante, Reporte

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Create your views here.

def index(request):
    return render(request, 'base.html')

def registrar_estudiante(request):
    if request.method == 'POST':
        Estudiante.objects.create(
            nombre=request.POST['nombre'],
            apellido=request.POST['apellido'],
            tipo_documento=request.POST['tipo_documento'],
            numero_documento=request.POST['numero_documento'],
            telefono=request.POST['telefono'],
            correo=request.POST['correo'],
            ficha=request.POST['ficha'],
            programa=request.POST['programa'],
            horario=request.POST['horario']
        )
        return redirect('registrar')
    return render(request, 'registrar_estudiante.html')

def reporte_academico(request):
    estudiantes = Estudiante.objects.all()
    if request.method == 'POST':
        estudiante_id = request.POST['estudiante']
        estudiante = Estudiante.objects.get(id=estudiante_id)
        Reporte.objects.create(
            estudiante=estudiante,
            tipo=request.POST['tipo'],
            detalle=request.POST['detalle'],
            instructor=request.POST['instructor']
        )
        return redirect('reporte')
    return render(request, 'reporte_academico.html', {'estudiantes': estudiantes})

def listado_estudiantes(request):
    estudiantes = Estudiante.objects.all()
    return render(request, 'listado_estudiantes.html',
                  {'estudiantes': estudiantes})

def listado_reportes(request):
    reporte = Reporte.objects.all()
    return render(request, 'listado_reportes.html',
                  {'reportes': reporte})

# Eliminar
def eliminar_estudiante(request, estudiante_id):
    estudiante = Estudiante.objects.get(id=estudiante_id)
    if request.method == 'POST':
        if 'confirmar' in request.POST:
            estudiante.delete()
            messages.success(request, 'Estudiante eliminado exitosamente')
            return redirect('listado_estudiante')
        else:
            return redirect('eliminar_estudiante')
    return render(request, 'confirmar_eliminacion.html', {'estudiante': estudiante})

def eliminar_reporte(request, reporte_id):
    reporte = Reporte.objects.get(id=reporte_id)
    if request.method == 'POST':
        if 'confirmar' in request.POST:
            reporte.delete()
            messages.success(request, 'Reporte eliminado exitosamente')
            return redirect('listado_reportes')
        else:
            return redirect('eliminar_reporte')
    return render(request, 'confirmar_eliminacionReporte.html', {'reporte': reporte})
''

# Editar
def editar_estudiante(request, estudiante_id):
    estudiante = Estudiante.objects.get(id=estudiante_id)
    if request.method == 'POST':
        try:
            estudiante.nombre = request.POST['nombre']
            estudiante.apellido = request.POST['apellido']
            estudiante.tipo_documento = request.POST['tipo_documento']
            estudiante.numero_documento = request.POST['numero_documento']
            estudiante.telefono = request.POST['telefono']
            estudiante.correo = request.POST['correo']
            estudiante.ficha = request.POST['ficha']
            estudiante.programa = request.POST['programa']
            estudiante.horario = request.POST['horario']
            estudiante.save()
            messages.success(request, 'Estudiante actualizado exitosamente')
        except:
            messages.error(request, 'Error al actualizar el estudiante')
        return redirect('listado_estudiante')
    return render(request, 'editar_estudiante.html', {'estudiante': estudiante})

def editar_reporte(request, reporte_id):
    reporte = Reporte.objects.get(id=reporte_id)
    if request.method == 'POST':
        try:
            reporte.estudiante = request.POST['estudiante']
            reporte.tipo = request.POST['tipo']
            reporte.detalle = request.POST['detalle']
            reporte.instructor = request.POST['instructor']
            reporte.save()
            messages.success(request, 'Reporte actualizado exitosamente')
        except:
            messages.error(request, 'Error al actualizar el reporte')
        return redirect('listado_reportes')
    return render(request, 'editar_reporte.html', {'reporte': reporte})


# Exportar a PDF
def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="estudiantes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    data = [['Nombre', 'Apellido', 'Tipo Doc.', 'N° Doc.', 'Teléfono', 'Correo', 'Ficha', 'Programa', 'Horario']]
    estudiantes = Estudiante.objects.all()

    for est in estudiantes:
        data.append([
            est.nombre, est.apellido, est.tipo_documento, est.numero_documento,
            est.telefono, est.correo, est.ficha, est.programa, est.horario
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003366")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey)
    ]))

    elements.append(table)
    doc.build(elements)
    return response


# Exportar a Excel
def exportar_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estudiantes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    headers = ['Nombre', 'Apellido', 'Tipo de Documento', 'Número de Documento',
               'Teléfono', 'Correo', 'Ficha', 'Programa', 'Horario']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')

    estudiantes = Estudiante.objects.all()
    for est in estudiantes:
        ws.append([
            est.nombre, est.apellido, est.tipo_documento, est.numero_documento,
            est.telefono, est.correo, est.ficha, est.programa, est.horario
        ])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(response)
    return response



def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reportes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    data = [['Estudiante', 'Tipo', 'Detalle', 'Instructor', 'Fecha']]
    reportes = Reporte.objects.all()

    for rep in reportes:
        data.append([
            f"{rep.estudiante.nombre} {rep.estudiante.apellido}",
            rep.tipo,
            rep.detalle,
            rep.instructor,
            rep.fecha.strftime('%Y-%m-%d')
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003366")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey)
    ]))

    elements.append(table)
    doc.build(elements)
    return response


# Exportar a Excel
def exportar_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reportes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reportes"

    headers = ['Estudiante', 'Tipo', 'Detalle', 'Instructor', 'Fecha']
    ws.append(headers)

    # Estilo de cabecera
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')

    reportes = Reporte.objects.all()
    for rep in reportes:
        ws.append([
            f"{rep.estudiante.nombre} {rep.estudiante.apellido}",
            rep.tipo,
            rep.detalle,
            rep.instructor,
            rep.fecha.strftime('%Y-%m-%d')
        ])

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(response)
    return response